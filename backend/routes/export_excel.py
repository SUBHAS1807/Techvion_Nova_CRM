"""
TechvionNova CRM - Professional Excel (.xlsx) Lead Report Generator
Generates multi-sheet workbook with branded styling, conditional status colors,
hyperlinks, auto-filters, freeze panes, and an executive KPI summary sheet.
"""

import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Machine website-status code -> human display label (worldwide collector)
_WS_CODE_LABELS = {
    "NO_WEBSITE": "No Website",
    "HAS_WEBSITE": "Has Website",
    "WEBSITE_INACCESSIBLE": "Website Inaccessible",
    "WEBSITE_UNKNOWN": "Unknown",
}


def _ws_label(lead) -> str:
    """Prefer the machine code's label; fall back to legacy human value."""
    code = getattr(lead, "website_status_code", None)
    if code and code in _WS_CODE_LABELS:
        return _WS_CODE_LABELS[code]
    return lead.website_status or "Unknown"


# Brand Color Palette (Hex)
NAVY_HEADER_FILL = "0F172A"      # Main title & brand fill
SLATE_HEADER_FILL = "1E293B"     # Column header fill
ZEBRA_FILL = "F8FAFC"            # Alternating row fill
WHITE_FILL = "FFFFFF"
BORDER_COLOR = "CBD5E1"          # Light border

# Status Fills and Font Colors (Soft Pill Badges)
STATUS_STYLES = {
    # Website Status
    "Good": {"fill": "DCFCE7", "font": "166534", "bold": True},
    "No Website": {"fill": "FEE2E2", "font": "991B1B", "bold": True},
    "Outdated": {"fill": "FFEDD5", "font": "9A3412", "bold": True},
    "Broken": {"fill": "FEE2E2", "font": "991B1B", "bold": True},
    "Under Construction": {"fill": "FEF3C7", "font": "92400E", "bold": True},
    "E-commerce Website": {"fill": "E0E7FF", "font": "3730A3", "bold": True},
    "Booking Website": {"fill": "F3E8FF", "font": "6B21A8", "bold": True},
    "Unknown": {"fill": "F1F5F9", "font": "475569", "bold": False},

    # Outreach Status
    "Not Contacted": {"fill": "F1F5F9", "font": "475569", "bold": False},
    "Contacted": {"fill": "DBEAFE", "font": "1E40AF", "bold": True},
    "Follow-up": {"fill": "FFEDD5", "font": "9A3412", "bold": True},
    "Completed": {"fill": "DCFCE7", "font": "166534", "bold": True},
    "Do Not Contact": {"fill": "FEE2E2", "font": "991B1B", "bold": True},

    # Deal Status
    "Open": {"fill": "DBEAFE", "font": "1E40AF", "bold": True},
    "Negotiation": {"fill": "FFEDD5", "font": "9A3412", "bold": True},
    "Won": {"fill": "DCFCE7", "font": "166534", "bold": True},
    "Lost": {"fill": "FEE2E2", "font": "991B1B", "bold": True},
    "On Hold": {"fill": "F1F5F9", "font": "475569", "bold": False},

    # Interested / Agreed
    "Interested": {"fill": "F3E8FF", "font": "6B21A8", "bold": True},
    "Agreed": {"fill": "DCFCE7", "font": "166534", "bold": True},
    "Pending": {"fill": "F1F5F9", "font": "64748B", "bold": False},
    "Rejected": {"fill": "FEE2E2", "font": "991B1B", "bold": True},
}


def create_excel_report(leads) -> bytes:
    """
    Generate professional Excel workbook for a list of Lead model instances.
    Returns binary bytes of the .xlsx file.
    """
    wb = Workbook()

    # ── SHEET 1: Leads ──────────────────────────────────────────────────
    ws_leads = wb.active
    ws_leads.title = "Leads"
    ws_leads.views.sheetView[0].showGridLines = True

    # Standard styling tokens
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # 1. Title Banner (Rows 1-2)
    ws_leads.merge_cells("A1:K1")
    title_cell = ws_leads["A1"]
    title_cell.value = "TECHVIONNOVA CRM — OFFICIAL LEADS EXPORT"
    title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=NAVY_HEADER_FILL, end_color=NAVY_HEADER_FILL, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_leads.row_dimensions[1].height = 36

    ws_leads.merge_cells("L1:AM1")
    meta_cell = ws_leads["L1"]
    meta_cell.value = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')} | Total Records: {len(leads)}"
    meta_cell.font = Font(name="Segoe UI", size=10, bold=False, color="94A3B8")
    meta_cell.fill = PatternFill(start_color=NAVY_HEADER_FILL, end_color=NAVY_HEADER_FILL, fill_type="solid")
    meta_cell.alignment = Alignment(horizontal="right", vertical="center")

    # Row 2 Blank padding
    ws_leads.row_dimensions[2].height = 6

    # 2. Header Columns (Row 3)
    headers = [
        "Marked",
        "Lead ID",
        "Business Name",
        "Owner Name",
        "Business Type",
        "City",
        "Country",
        "Region",
        "State / Province",
        "Postal Code",
        "Lead Source",
        "Phone",
        "Email",
        "Email Source",
        "Email Verification",
        "Current Website",
        "Instagram",
        "Facebook",
        "Website Status",
        "Preferred Contact Channel",
        "First Contact Date",
        "Outreach Status",
        "Response Status",
        "Interested / Agreed",
        "Website Requirement",
        "Estimated Budget",
        "Currency",
        "Proposal Status",
        "Deal Status",
        "Project Status",
        "Next Follow-up Date",
        "Remarks",
        "Google Place ID",
        "Google Maps URL",
        "Address",
        "Google Rating",
        "Google Reviews",
        "Lead Score",
        "Created At",
    ]

    ws_leads.row_dimensions[3].height = 28
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws_leads.cell(row=3, column=col_idx, value=header_text)
        cell.font = Font(name="Segoe UI", size=10.5, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=SLATE_HEADER_FILL, end_color=SLATE_HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border

    # 3. Populate Lead Data Rows
    current_row = 4
    for idx, l in enumerate(leads):
        row_fill = PatternFill(
            start_color=ZEBRA_FILL if idx % 2 == 1 else WHITE_FILL,
            end_color=ZEBRA_FILL if idx % 2 == 1 else WHITE_FILL,
            fill_type="solid"
        )

        first_contact_str = l.first_contact_date.strftime("%Y-%m-%d") if l.first_contact_date else ""
        next_followup_str = l.next_followup_date.strftime("%Y-%m-%d") if l.next_followup_date else ""
        created_at_str = l.created_at.strftime("%Y-%m-%d %H:%M") if l.created_at else ""

        row_data = [
            "Yes" if l.is_marked else "No",
            l.lead_id,
            l.business_name or "",
            l.owner_name or "Unknown",
            l.business_type or "",
            l.city or "",
            l.country or "",
            l.region or "",
            l.state_province or l.region or "",
            l.postal_code or "",
            l.lead_source or "Google Places API",
            l.phone or "",
            l.email or "",
            getattr(l, "email_source", "") or ("Business Website" if l.email else ""),
            getattr(l, "email_status", "") or getattr(l, "email_verification_status", "") or "Not Checked",
            l.current_website or "",
            l.instagram or "",
            l.facebook or "",
            _ws_label(l),
            l.preferred_contact_channel or "",
            first_contact_str,
            l.outreach_status or "Not Contacted",
            l.response_status or "No Response",
            l.interested_agreed or "Pending",
            l.website_requirement or "",
            l.estimated_budget or "",
            l.currency or "",
            l.proposal_status or "Not Sent",
            l.deal_status or "Open",
            l.project_status or "Not Started",
            next_followup_str,
            l.remarks or "",
            l.google_place_id or "",
            l.google_maps_url or l.source_url or "",
            l.address or "",
            float(l.rating or l.google_rating) if (l.rating or l.google_rating) is not None else "",
            int(l.review_count or l.google_reviews) if (l.review_count or l.google_reviews) is not None else "",
            int(l.lead_score or 0),
            created_at_str,
        ]

        ws_leads.row_dimensions[current_row].height = 22

        for col_idx, val in enumerate(row_data, 1):
            cell = ws_leads.cell(row=current_row, column=col_idx)
            header_name = headers[col_idx - 1]
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = Font(name="Segoe UI", size=10, color="1E293B")
            cell.alignment = Alignment(vertical="center", horizontal="left")

            # Column specific formatting
            if header_name in ("Marked", "Lead ID", "Google Rating", "Google Reviews", "Lead Score", "First Contact Date", "Next Follow-up Date", "Created At"):
                cell.alignment = Alignment(vertical="center", horizontal="center")

            # Phone numbers formatted strictly as string
            if header_name == "Phone":
                cell.number_format = "@"
                cell.value = str(val) if val else ""
                cell.alignment = Alignment(vertical="center", horizontal="left")

            # Email hyperlink
            elif header_name == "Email" and val:
                cell.value = val
                cell.font = Font(name="Segoe UI", size=10, color="2563EB", underline="single")
                cell.hyperlink = f"mailto:{val}"

            # Website hyperlink
            elif header_name in ("Current Website", "Google Maps URL") and val:
                clean_url = val if str(val).startswith("http") else f"https://{val}"
                cell.value = str(val)
                cell.font = Font(name="Segoe UI", size=10, color="2563EB", underline="single")
                try:
                    cell.hyperlink = clean_url
                except Exception:
                    pass

            # Status pill styling
            elif header_name in ("Website Status", "Outreach Status", "Deal Status", "Interested / Agreed") and val:
                cell.value = str(val)
                cell.alignment = Alignment(vertical="center", horizontal="center")
                style = STATUS_STYLES.get(str(val))
                if style:
                    cell.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
                    cell.font = Font(name="Segoe UI", size=9.5, bold=style.get("bold", False), color=style["font"])
            else:
                cell.value = val

        current_row += 1

    # Freeze panes below header row (at row 4)
    ws_leads.freeze_panes = "A4"

    # Auto filter on header row
    last_col_letter = get_column_letter(len(headers))
    ws_leads.auto_filter.ref = f"A3:{last_col_letter}{current_row - 1}"

    # Auto-fit column widths with +3 padding
    for col in ws_leads.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip title banner rows when calculating width
            if cell.row < 3:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_leads.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 48)


    # ── SHEET 2: Summary ────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.views.sheetView[0].showGridLines = True

    # 1. Summary Title Banner
    ws_summary.merge_cells("A1:F1")
    s_title = ws_summary["A1"]
    s_title.value = "TECHVIONNOVA CRM — EXECUTIVE SUMMARY & METRICS"
    s_title.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    s_title.fill = PatternFill(start_color=NAVY_HEADER_FILL, end_color=NAVY_HEADER_FILL, fill_type="solid")
    s_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_summary.row_dimensions[1].height = 36

    # 2. Compute Real Aggregate Stats from the Exported Leads
    total_leads = len(leads)
    leads_with_web = sum(1 for l in leads if l.current_website and l.current_website.strip())
    leads_no_web = total_leads - leads_with_web
    emails_found = sum(1 for l in leads if l.email and l.email.strip())
    emails_missing = total_leads - emails_found
    contacted_count = sum(1 for l in leads if l.outreach_status in ("Contacted", "Follow-up", "Completed"))
    interested_count = sum(1 for l in leads if l.interested_agreed in ("Interested", "Agreed") or l.response_status in ("Interested", "Positive"))
    proposals_sent = sum(1 for l in leads if l.proposal_status in ("Sent", "Accepted", "Revised"))
    deals_won = sum(1 for l in leads if l.deal_status == "Won")
    deals_lost = sum(1 for l in leads if l.deal_status == "Lost")

    # Estimated budget sum (clean non-digits)
    total_budget_val = 0
    for l in leads:
        if l.estimated_budget:
            nums = re.findall(r"\d+", str(l.estimated_budget).replace(",", ""))
            if nums:
                total_budget_val += int(nums[0])

    kpi_metrics = [
        ("Total Leads Exported", total_leads, "Total business records in this export batch"),
        ("Leads With Website", leads_with_web, f"{round((leads_with_web / total_leads * 100) if total_leads else 0, 1)}% of total leads"),
        ("Leads Without Website", leads_no_web, f"{round((leads_no_web / total_leads * 100) if total_leads else 0, 1)}% high-priority website sales targets"),
        ("Public Emails Found", emails_found, f"{round((emails_found / total_leads * 100) if total_leads else 0, 1)}% contact rate via business email"),
        ("Emails Missing", emails_missing, "Requires phone outreach or website analysis"),
        ("Contacted Leads", contacted_count, "Leads moved past initial cold stage"),
        ("Interested / Qualified", interested_count, "Positive response or agreed to proposal"),
        ("Proposals Sent", proposals_sent, "Formal proposals delivered"),
        ("Deals Won", deals_won, "Successfully closed customer contracts"),
        ("Deals Lost", deals_lost, "Archived / rejected opportunities"),
    ]

    # Section 1 Header
    ws_summary.cell(row=3, column=1, value="CRM Pipeline Key Performance Indicators").font = Font(name="Segoe UI", size=12, bold=True, color=NAVY_HEADER_FILL)
    ws_summary.row_dimensions[3].height = 24

    ws_summary.cell(row=4, column=1, value="Metric Name").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary.cell(row=4, column=1).fill = PatternFill(start_color=SLATE_HEADER_FILL, end_color=SLATE_HEADER_FILL, fill_type="solid")
    ws_summary.cell(row=4, column=1).border = thin_border

    ws_summary.cell(row=4, column=2, value="Count / Value").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary.cell(row=4, column=2).fill = PatternFill(start_color=SLATE_HEADER_FILL, end_color=SLATE_HEADER_FILL, fill_type="solid")
    ws_summary.cell(row=4, column=2).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=4, column=2).border = thin_border

    ws_summary.cell(row=4, column=3, value="Description / Insights").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary.cell(row=4, column=3).fill = PatternFill(start_color=SLATE_HEADER_FILL, end_color=SLATE_HEADER_FILL, fill_type="solid")
    ws_summary.cell(row=4, column=3).border = thin_border

    s_row = 5
    for m_name, m_val, m_desc in kpi_metrics:
        ws_summary.row_dimensions[s_row].height = 20
        c1 = ws_summary.cell(row=s_row, column=1, value=m_name)
        c2 = ws_summary.cell(row=s_row, column=2, value=m_val)
        c3 = ws_summary.cell(row=s_row, column=3, value=m_desc)

        c1.font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
        c2.font = Font(name="Segoe UI", size=10.5, bold=True, color="0284C7" if "Won" in m_name or "Found" in m_name else "0F172A")
        c2.alignment = Alignment(horizontal="center")
        c3.font = Font(name="Segoe UI", size=9.5, color="64748B")

        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border

        r_fill = PatternFill(start_color=ZEBRA_FILL if s_row % 2 == 1 else WHITE_FILL, fill_type="solid")
        c1.fill = r_fill
        c2.fill = r_fill
        c3.fill = r_fill
        s_row += 1

    # Section 2: Website Status Breakdown
    s_row += 2
    ws_summary.cell(row=s_row, column=1, value="Website Quality & Status Breakdown").font = Font(name="Segoe UI", size=12, bold=True, color=NAVY_HEADER_FILL)
    s_row += 1

    ws_summary.cell(row=s_row, column=1, value="Website Status").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary.cell(row=s_row, column=1).fill = PatternFill(start_color=SLATE_HEADER_FILL, end_color=SLATE_HEADER_FILL, fill_type="solid")
    ws_summary.cell(row=s_row, column=2, value="Lead Count").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary.cell(row=s_row, column=2).fill = PatternFill(start_color=SLATE_HEADER_FILL, end_color=SLATE_HEADER_FILL, fill_type="solid")
    ws_summary.cell(row=s_row, column=2).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=s_row, column=3, value="Share of Leads").font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws_summary.cell(row=s_row, column=3).fill = PatternFill(start_color=SLATE_HEADER_FILL, end_color=SLATE_HEADER_FILL, fill_type="solid")

    web_statuses = ["No Website", "Outdated", "Good", "Broken", "Under Construction", "E-commerce Website", "Booking Website", "Unknown"]
    s_row += 1
    for st in web_statuses:
        count = sum(1 for l in leads if (l.website_status or "Unknown") == st)
        if count == 0 and st not in ("No Website", "Outdated", "Good"):
            continue
        c1 = ws_summary.cell(row=s_row, column=1, value=st)
        c2 = ws_summary.cell(row=s_row, column=2, value=count)
        pct = f"{round((count / total_leads * 100) if total_leads else 0, 1)}%"
        c3 = ws_summary.cell(row=s_row, column=3, value=pct)

        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border
        c2.alignment = Alignment(horizontal="center")
        c3.alignment = Alignment(horizontal="center")

        st_style = STATUS_STYLES.get(st)
        if st_style:
            c1.fill = PatternFill(start_color=st_style["fill"], fill_type="solid")
            c1.font = Font(name="Segoe UI", size=9.5, bold=True, color=st_style["font"])
        s_row += 1

    # Auto column widths for summary sheet
    ws_summary.column_dimensions["A"].width = 34
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 46
    ws_summary.column_dimensions["D"].width = 20

    # Save to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
